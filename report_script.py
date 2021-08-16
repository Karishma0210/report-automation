print("Python Script is running")
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

in_filename = "C:\\Users\\karishma\\Documents\\jNoteBook\\msh\\Source Data.xlsm"
out_filename = "C:\\Users\\karishma\\Documents\\jNoteBook\\msh\\Report Template.xlsx"

# in_filename = "Source Data.xlsm"
# out_filename = "Report Template.xlsx"

src_df = pd.read_excel(in_filename)

wb = load_workbook(out_filename)
report_ws = wb.worksheets[0]
print("Python Script is running")
#total payment (unpaid, paid both)
total_pay = src_df['Claimed Amount'].sum()
# print("total pay:", total_pay)
cell = report_ws.cell(3, 3)
cell.value = total_pay
cell.number_format = '#,##0.00'

#unpaid payment - one without the Payment Date
unpaid_total = src_df[src_df['Payment Date'].isnull()]['Claimed Amount'].sum()
# print("unpaid (without the dates):", unpaid_total)
cell = report_ws.cell(4, 3)
cell.value = unpaid_total
cell.number_format = '#,##0.00'

#convert date to date type
src_df['Date of Birth'] = src_df['Date of Birth'].astype('datetime64')

#add age column
now = pd.Timestamp('now')
src_df['Age'] = now - src_df['Date of Birth']
src_df['Age'] = src_df['Age'].astype('timedelta64[D]')/365.25
# print(src_df['Age'])

#Population count by Age Group & Gender
age_bin = [0, 16, 26, 36, 51, 65, 125]
src_df['Age_Group'] = pd.cut(src_df['Age'], age_bin, labels=['0-15', '16-25', '26-35', '36-50', '51-65', '65+'])
pop_count_by_age_group_gender = src_df.groupby(['Age_Group', 'Gender'])['Member'].count()
pop_count_by_age_group_gender = pop_count_by_age_group_gender.reset_index().sort_values(by='Gender', ascending=False).reset_index(drop=True)
# print(type(pop_count_by_age_group_gender))
# print("Population count by age group & gender\n", pop_count_by_age_group_gender)

#put record values in cells
for row in range(6,8):
    for col in range(3,9):
        cell = report_ws.cell(row, col)
        if row==6:
            cell.value = pop_count_by_age_group_gender.iloc[col-3]['Member']
            
        else:
            cell.value = pop_count_by_age_group_gender.iloc[col-3 + 6]['Member']
            
        cell.number_format = '#,##0'
            


#Paid amount by Member Type
paid_by_member_type = src_df[src_df['Payment Date'].notnull()].groupby('Member Type')['Claimed Amount'].sum()
total_paid = paid_by_member_type.sum()
# print("Paid by Member Type\n", paid_by_member_type)
for row in range(9,12):
    cell = report_ws.cell(row, 3)
    cell.value = paid_by_member_type[report_ws.cell(row, 2).value]
    cell.number_format = '#,##0.00'

# print("total paid by all members:", total_paid)
cell = report_ws.cell(12, 3)
cell.value = total_paid
cell.number_format = '#,##0.00'

#Paid by Network
paid_by_network = src_df[src_df['Payment Date'].notnull()].groupby('Network')['Claimed Amount'].sum()
# print("Paid by Network:\n", paid_by_network)
for row in range(14, 16):
    cell = report_ws.cell(row, 3)
    cell.value = paid_by_network[report_ws.cell(row, 2).value]
    cell.number_format = '#,##0.00'

#Paid amount by Month/ Year
src_df['Payment Date'] = src_df['Payment Date'].astype('datetime64')
src_df['payment_month'] = src_df['Payment Date'].dt.month.astype('Int64')
src_df['payment_year'] = src_df['Payment Date'].dt.year.astype('Int64')
# src_df['payment_month_year'] = src_df['Payment Date'].dt.strftime('%m-%Y')

value_by_month = src_df[src_df['payment_month'].notnull()].reset_index(drop=True).groupby('payment_month')['Claimed Amount'].sum()
# print("Paid value by month", value_by_month)

for row in range(17, 29):
#     print()
    cell = report_ws.cell(row, 6)
    cell.value = value_by_month[row-16]
    cell.number_format = '#,##0.00'
    
#get minimum and maximum range year by month
paid_by_year_range = src_df[src_df['payment_month'].notnull()]
paid_by_year_range = paid_by_year_range.groupby('payment_month').agg({'payment_year': ['min', 'max']})

paid_by_year_range['year_range'] = paid_by_year_range['payment_year'][['min','max']].astype('Int64').astype(str).apply(lambda x: '-'.join(x), axis=1)

for row in range(17, 29):
#     print()
    cell = report_ws.cell(row, 3)
    cell.value = paid_by_year_range['year_range'][row-16]

# src_df
# print(src_df.dtypes)

#save all the values to the workbook
wb.save(out_filename)
wb.close()
print("Operations done")